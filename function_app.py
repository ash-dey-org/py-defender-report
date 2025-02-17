# Extract report from Defender API
# upload to Sumologic
# download file from sharepoint
# update excel
# upload updated to sharepoint
# send email

# pip install azure-identity
# pip install azure-communication-identity
# pip install azure-communication-email
# pip install azure-keyvault-secrets
# pip install azure-keyvault-certificates
# pip install office365-REST-Python-Client
# pip install openpyxl
# pip install azure-functions

import os
import requests
import time
import json
import csv
import urllib.request
import urllib.parse
from datetime import datetime, timezone
import logging
import base64

import pandas as pd
from openpyxl import load_workbook

import azure.functions as func
from azure.communication.email import EmailClient
from azure.identity import DefaultAzureCredential
from azure.keyvault.secrets import SecretClient
from azure.keyvault.certificates import CertificateClient

from office365.sharepoint.client_context import ClientContext



# define environment variables and other fixed parameters
tenantId = os.environ.get('azure_tenant_id')
appId = os.environ.get('defender_app_id')
appSecret = os.environ.get('defender_app_secret')
sumoUrl = os.environ.get('sumo_collector_url')
kvUrl = os.environ.get('kv_url')
certName = os.environ.get('cert_name')
sharepointUrl = os.environ.get('sharepoint_url')
sharepointDir = os.environ.get('sharepoint_dir')
sharepointFilePath = os.environ.get('sharepoint_file_path')
sharepointFilePathSummary = os.environ.get('sharepoint_file_path_summary')
sharepointFilePathCritical = os.environ.get('sharepoint_file_path_critical')
sharepointFilePathHigh = os.environ.get('sharepoint_file_path_high')
sharepointFilePathMedium = os.environ.get('sharepoint_file_path_medium')
sharepointFilePathLow = os.environ.get('sharepoint_file_path_low')

urlauth = "https://login.microsoftonline.com/%s/oauth2/token" % (tenantId)

resourceAppIdUri = 'https://api.securitycenter.microsoft.com'

body = {
    'resource' : resourceAppIdUri,
    'client_id' : appId,
    'client_secret' : appSecret,
    'grant_type' : 'client_credentials'
}

# Get current date in UTC
todays_date = datetime.now(timezone.utc)


# define function app and its scehdule
app = func.FunctionApp()

'''
@app.function_name(name="updateSharepoint")
@app.schedule(schedule="0 0 */3 * * *",
# @app.schedule(schedule="0 */10 * * * *",
              arg_name="updateSharepoint",
              run_on_startup=False)
def test_function1(updateSharepoint: func.TimerRequest) -> None:
    utc_timestamp = datetime.datetime.now(datetime.timezone.utc).isoformat()
    if updateSharepoint.past_due:
        logging.info('The timer is past due!')
    logging.info('Python timer trigger function ran at %s', utc_timestamp)
    result = defenderReport()
    # uploadDataToSumologic(result[0],sumoUrl)
    print("txt file name is:", result[0])
    print("csv file name is:", result[1])
    ctx = authneticateToSharepoint(kvUrl, certName, sharepointUrl)
    fileDownload = downloadFile(ctx, sharepointFilePath)
    updateExcel(fileDownload, result[1])
    uploadFile (ctx, fileDownload, sharepointDir)
    sendEmail(result[1])
'''

@app.function_name(name="updateSharepoint")
@app.schedule(schedule="0 20 6 * * *",
# @app.schedule(schedule="0 */10 * * * *",
              arg_name="updateSharepoint",
              run_on_startup=True)
def test_function1(updateSharepoint: func.TimerRequest) -> None:
    utc_timestamp = datetime.datetime.now(datetime.timezone.utc).isoformat()
    if updateSharepoint.past_due:
        logging.info('The timer is past due!')
    logging.info('Python timer trigger function ran at %s', utc_timestamp)
    # extract summary report
    summary_file = defenderSummary("InfraReporting-defender-summary")
    print(summary_file)

    # Extract device info
    device_file=defenderAdvanceHunting("device")
    print(device_file)

    # Extract evidence info
    evidence_file=defenderAdvanceHunting("evidence")
    print(evidence_file)

    # Exctract Vulerability reports
    critical_file=defenderAdvanceHunting("critical")
    print(critical_file)
    high_file=defenderAdvanceHunting("high")
    print(high_file)
    medium_file=defenderAdvanceHunting("medium")
    print(medium_file)
    low_file=defenderAdvanceHunting("low")
    print(low_file)

    # Merge vulnerability files with last see time stamp
    critical_file_merged =  mergeFiles(device_file,critical_file,evidence_file)
    print(critical_file_merged)
    high_file_merged =  mergeFiles(device_file,high_file,evidence_file)
    print(high_file_merged)
    medium_file_merged =  mergeFiles(device_file,medium_file,evidence_file)
    print(medium_file_merged)
    low_file_merged =  mergeFiles(device_file,low_file,evidence_file)
    print(low_file_merged)

    ctx = authneticateToSharepoint(kvUrl, certName, sharepointUrl)

    uploadFile (ctx, summary_file, sharepointDir)
    uploadFile (ctx, evidence_file, sharepointDir)
    uploadFile (ctx, critical_file_merged, sharepointDir)
    uploadFile (ctx, high_file_merged, sharepointDir)
    uploadFile (ctx, medium_file_merged, sharepointDir)
    uploadFile (ctx, low_file_merged, sharepointDir)

@app.function_name(name="updateSumologic")
@app.schedule(schedule="0 15 22 * * *",
# @app.schedule(schedule="0 */10 * * * *",
              arg_name="updateSumologic",
              run_on_startup=False)
def test_function2(updateSumologic: func.TimerRequest) -> None:
    utc_timestamp = datetime.datetime.now(datetime.timezone.utc).isoformat()
    if updateSumologic.past_due:
        logging.info('The timer is past due!')
    logging.info('Python timer trigger function ran at %s', utc_timestamp)
    result = defenderReport()
    uploadDataToSumologic(result[0],sumoUrl)
    print("txt file name is:", result[0])
    print("csv file name is:", result[1])
    # ctx = authneticateToSharepoint(kvUrl, certName, sharepointUrl)
    # fileDownload = downloadFile(ctx, sharepointFilePath)
    # updateExcel(fileDownload, result[1])
    # uploadFile (ctx, fileDownload, sharepointDir)
    sendEmail(result[1])

def defenderReport():

    # Authenticate to Azure uisng aap client and secret and get token
    data = urllib.parse.urlencode(body).encode("utf-8")

    req = urllib.request.Request(url, data)
    response = urllib.request.urlopen(req)
    jsonResponse = json.loads(response.read())
    aadToken = jsonResponse["access_token"]

    # run query using the token
    queryFile = open("query.txt", 'r')
    query = queryFile.read()
    queryFile.close()
    # print(query)


    secUrl = "https://api.securitycenter.microsoft.com/api/advancedqueries/run"
    headers = {
        'Content-Type' : 'application/json',
        'Accept' : 'application/json',
        'Authorization' : "Bearer " + aadToken
    }

    data = json.dumps({ 'Query' : query }).encode("utf-8")
    # print(data)
    req = urllib.request.Request(secUrl, data, headers)
    response = urllib.request.urlopen(req)
    # print("getting data from Defender portal, can take appox 5 sec....")
    # time.sleep(5)
    jsonResponse = json.loads(response.read())
    schema = jsonResponse["Schema"]
    results = jsonResponse["Results"]
    # print(results)

    # export output from query to a text file containing json array
    """
    home_directory = os.path.expanduser("~")
    downloads_folder = os.path.join(home_directory, "Downloads")
    if not os.path.exists(downloads_folder):
        os.makedirs(downloads_folder)

    """

    current_date = time.strftime("%Y%m%d")
    directory_path = "/tmp/"
    file_path_txt = f"{directory_path}defender_log_{current_date}.txt"
    file_path_csv = f"{directory_path}defender_log_{current_date}.csv"
    # filenametxt = f"defender_log_{current_date}.txt"
    # filenamecsv = f"defender_log_{current_date}.csv"
    # file_path_txt = os.path.join(downloads_folder, filenametxt)
    # file_path_csv = os.path.join(downloads_folder, filenamecsv)


    # write to csv file
    print("writing output to csv file......")
    outputFile = open(file_path_csv, 'w')
    output = csv.writer(outputFile)
    output.writerow(results[0].keys())
    for result in results:
        output.writerow(result.values())
    outputFile.close()


    # write output to text file
    print("writing output to txt file......")
    with open(file_path_txt, "w", encoding="utf-8") as output_file:
        for result in results:
            result_json = json.dumps(result, ensure_ascii=False)
            output_file.write(result_json + "\n")
    output_file.close()
    # time.sleep(1)

    return file_path_txt, file_path_csv

def uploadDataToSumologic(file_path_txt:str,sumourl:str):
    # upload logs to Sumologic
    print("uploading data to sumologic......")
    cmd = 'curl -v -X POST -H "X-Sumo-Category:security/defender/hunting" -H "X-Sumo-Name:%s" -T %s %s --ssl-no-revoke' %(file_path_txt, file_path_txt, sumourl)
    # print(cmd)
    returned_value = os.system(cmd)
    # print('returned value:', returned_value)
    print("Done.... Check Somologic portal for uploaded data.")

def authneticateToSharepoint(kv_url:str, cert_name:str, sharepoint_url:str):

    # Securely retrieve secrets from Azure Key Vault
    credential = DefaultAzureCredential()
    certificate_client = CertificateClient(vault_url=kv_url, credential=credential)
    certificate = certificate_client.get_certificate(certificate_name=cert_name)
    client = SecretClient(vault_url=kv_url, credential=credential)

    # download pem file from key vault
    with open("temp.pem", "w") as pem_file:
        pem_file.write(client.get_secret(cert_name).value)

    client_id = appId   # client.get_secret("defender-app-id").value
    tenant_id = tenantId  # client.get_secret("azure-tenant-id").value
    cert_thumbprint = certificate.properties.x509_thumbprint.hex()

    cert_credentials = {
        "tenant": tenant_id,
        "client_id": client_id,
        "thumbprint": cert_thumbprint,
        "cert_path": "{0}/temp.pem".format(os.path.dirname(__file__)),
    }

    ctx = ClientContext(sharepoint_url).with_client_certificate(**cert_credentials)

    os.remove("temp.pem")
    return ctx

def downloadFile(ctx, sharepoint_file_path:str):
    download_path = os.path.join(os.getcwd(), os.path.basename(sharepoint_file_path))
    with open(download_path, "wb") as local_file:
        file = (
            ctx.web.get_file_by_server_relative_url(sharepoint_file_path)
            .download(local_file)
            .execute_query()
        )
    print("[Ok] file has been downloaded into: {0}".format(download_path))
    return download_path

def updateExcel(src_file:str, csv_file:str):
    # load workbook
    wb1 = load_workbook(src_file)

    # delete exisitng worksheet
    wb1.remove(wb1['Defender-critical-high-with-exp'])

    # create new worksheet at the end
    wb1.create_sheet("Defender-critical-high-with-exp")

    #load worksheet
    ws1 = wb1["Defender-critical-high-with-exp"]

     # open the csv file

    with open(csv_file) as f:
        reader = csv.reader(f, delimiter=',')

        for row in reader:
            ws1.append(row)

    # save worksheet
    wb1.save(src_file)
    print("Excel sheet ", src_file, "updated with new data")

def uploadFile(ctx,filename:str, sharepoint_dir:str):
    web = ctx.web.get_folder_by_server_relative_url(sharepoint_dir)

    with open(filename, "rb") as content_file:
        file_content = content_file.read()
    file = web.upload_file(os.path.basename(filename), file_content).execute_query()
    print("File has been uploaded into: {0}".format(file.serverRelativeUrl))

def sendEmail(attachment):

    connection_string = os.environ.get('comm_service_conn_string')

    with open(attachment, "rb") as file:
        file_bytes_b64 = base64.b64encode(file.read())

    message = {
        "content": {
            "subject": "Defender Vulnerability Report",
            "plainText": "Copy the content of the attached CSV file into the excel file in Teams channel: IT Infrastructure and Operations Team > Files > Security Reporting > Defender > Defender-critical-high-with-exploit.xlsx > 'Defender-critical-high-with-exp' Tab. After copying go to 'summary' tab > click anywhere inside table > PivotTable > Refress All data"

        },
        "recipients": {
            "to": [
                {
                    "address": "ash.dey@standards.org.au",
                    "displayName": "Ash Dey"
                }
            ]
        },
        "senderAddress": "DoNotReply@d737b976-8c58-45e0-a0e7-4c16f5c097c4.azurecomm.net",
        "replyTo": [
            {
                "address": "ash.dey@standards.org.au",  # Email address. Required.
                "displayName": "Ash Dey"  # Optional. Email display name.
            }
        ],
        "attachments": [
            {
                "name": "defender-report.csv",
                "contentType": "text/csv",
                "contentInBase64": file_bytes_b64.decode()
            }
        ]
    }


    POLLER_WAIT_TIME = 10

    try:
        # endpoint = "https://central-communication-service.australia.communication.azure.com"
        # email_client = EmailClient(endpoint, DefaultAzureCredential())
        print("set email client...")
        email_client = EmailClient.from_connection_string(connection_string)

        print("send email....")
        poller = email_client.begin_send(message);

        time_elapsed = 0
        while not poller.done():
            print("Email send poller status: " + poller.status())

            poller.wait(POLLER_WAIT_TIME)
            time_elapsed += POLLER_WAIT_TIME

            if time_elapsed > 18 * POLLER_WAIT_TIME:
                raise RuntimeError("Polling timed out.")

        if poller.result()["status"] == "Succeeded":
            print(f"Successfully sent the email (operation id: {poller.result()['id']})")
        else:
            raise RuntimeError(str(poller.result()["error"]))

    except Exception as ex:
        print(ex)

# Function to parse "publishedOn" with different formats
def parse_published_date(date_str):
    formats = [
        "%Y-%m-%dT%H:%M:%SZ",          # Format 1: 2024-11-28T00:00:00Z
        "%Y-%m-%dT%H:%M:%S.%fZ"        # Format 2: 2025-02-07T10:15:21.287Z
    ]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None  # Return None if parsing fails

# Function to extract summary report
def defenderSummary(output_file):

    # Authenticate to Azure uisng aap client and secret and get token
    data = urllib.parse.urlencode(body).encode("utf-8")

    req = urllib.request.Request(urlauth, data)
    response = urllib.request.urlopen(req)
    jsonResponse = json.loads(response.read())
    aadToken = jsonResponse["access_token"]


    # API endpoint
    url = "https://api.securitycenter.microsoft.com/api/Vulnerabilities"

    # Headers for authentication
    headers = {
            'Content-Type' : 'application/json',
            'Accept' : 'application/json',
            'Authorization' : "Bearer " + aadToken
        }

    # Make the API request
    response = requests.get(url, headers=headers)

    # Check if request was successful
    if response.status_code == 200:
        data = response.json()

        # Ensure data is a list of dictionaries
        if isinstance(data, dict) and "value" in data:
            # Extract the 'value' list
            vulnerabilities = data["value"]

    # Process each vulnerability
    filtered_vulnerabilities = []
    for vulnerability in vulnerabilities:
        if "name" in vulnerability:
            vulnerability["CveId"] = vulnerability.pop("name")  # Rename "name" to "CveId"

        # Check if CveId starts with "CVE-"
        if not vulnerability["CveId"].startswith("CVE-"):
            continue  # Skip this row

        if "description" in vulnerability:
            vulnerability["description"] = vulnerability["description"][:512]  # Limit description to 512 chars

        # Calculate "Age Days"
        days_ago = "N/A"
        if "publishedOn" in vulnerability:
            published_date = parse_published_date(vulnerability["publishedOn"])
            if published_date:
                days_ago = (todays_date - published_date).days

        vulnerability["Age (Days)"] = days_ago

        # Append valid entries to the list
        filtered_vulnerabilities.append(vulnerability)

    # Define CSV file name
    # csv_file = "vulnerabilities-summary.csv"
    directory_path = "/tmp/"
    file_path_csv = f"{directory_path}{output_file}.csv"

    # Check if there are valid entries
    if filtered_vulnerabilities:
        # Get CSV headers (including new "Day Ago" column)
        headers = list(filtered_vulnerabilities[0].keys())

        # Write to CSV file
        with open(file_path_csv, mode="w", newline="", encoding="utf-8") as file:
            writer = csv.DictWriter(file, fieldnames=headers, quoting=csv.QUOTE_ALL)

            # Write headers
            writer.writeheader()

            # Write data rows
            writer.writerows(filtered_vulnerabilities)

        print(f"CSV file '{file_path_csv}' has been created successfully with {len(filtered_vulnerabilities)} valid rows.")
    else:
        print("No valid CVE entries found. CSV file not created.")

    return file_path_csv

# Function to run defender advance queries
def defenderAdvanceHunting(query_file):

    # Authenticate to Azure uisng aap client and secret and get token
    data = urllib.parse.urlencode(body).encode("utf-8")

    req = urllib.request.Request(urlauth, data)
    response = urllib.request.urlopen(req)
    jsonResponse = json.loads(response.read())
    aadToken = jsonResponse["access_token"]

    # run query using the token
    queryFile = open(f"{query_file}", 'r')
    query = queryFile.read()
    queryFile.close()
    # print(query)


    secUrl = "https://api.securitycenter.microsoft.com/api/advancedqueries/run"
    headers = {
        'Content-Type' : 'application/json',
        'Accept' : 'application/json',
        'Authorization' : "Bearer " + aadToken
    }

    data = json.dumps({ 'Query' : query }).encode("utf-8")
    # print(data)
    req = urllib.request.Request(secUrl, data, headers)
    response = urllib.request.urlopen(req)
    # print("getting data from Defender portal, can take appox 5 sec....")
    # time.sleep(5)
    jsonResponse = json.loads(response.read())
    schema = jsonResponse["Schema"]
    results = jsonResponse["Results"]
    # print(results)

    # current_date = time.strftime("%Y%m%d")
    directory_path = "/tmp/"
    file_path_csv = f"{directory_path}InfraReporting-defender-{query_file}.csv"

    # write to csv file
    print("writing output to csv file......")
    outputFile = open(file_path_csv, 'w')
    output = csv.writer(outputFile)
    output.writerow(results[0].keys())
    for result in results:
        output.writerow(result.values())
    outputFile.close()

    return file_path_csv

# Function to merge files add device last seen dates and a few parameters to vulnerability files
def mergeFiles(device_file,vulnerability_file, evidence_file):

    output_file = vulnerability_file.replace(".csv", "-merged.csv")

    # Read the CSV files
    device_df = pd.read_csv(device_file)
    vul_df = pd.read_csv(vulnerability_file)
    # evidence_df = pd.read_csv(evidence_file)

    # Select important columns from the device file
    important_columns_dev = ["DeviceId", "LatestEntry", "PublicIP", "LoggedOnUsers", "AzureResourceId"]
    device_df = device_df[important_columns_dev]
    important_columns_vul = ["DeviceId", "DeviceName", "OSPlatform", "OSVersion", "SoftwareVendor", "SoftwareName", "SoftwareVersion", "CveId", "VulnerabilitySeverityLevel", "RecommendedSecurityUpdate"]
    vul_df = vul_df[important_columns_vul]
    # important_columns_evidence = ["DeviceId", "DiskPaths", "RegistryPaths"]
    # evidence_df = evidence_df[important_columns_evidence]

    # Convert LatestEntry to datetime and compute days since last seen
    device_df["LatestEntry"] = pd.to_datetime(device_df["LatestEntry"], errors="coerce")
    today = datetime.now(timezone.utc)
    device_df["DeviceSeenDaysAgo"] = (today - device_df["LatestEntry"]).dt.days

    # Merge: Left join on "DeviceId" to keep all rows from the critical file
    merged_df = vul_df.merge(device_df, on="DeviceId", how="left")

    # Merge evidence data into merged_df
    # merged_df = merged_df.merge(evidence_df, on="DeviceId", how="left")

    # Save the updated file
    merged_df.to_csv(output_file, index=False)

    print(f"Updated file saved as {output_file}")
    return  output_file

'''
# this part goes inside test function at the top
# used for local testing, comment this section after testing
# uncomment top section and # import azure.functions as func - before pushing code into azure function app
result = defenderReport()
uploadDataToSumologic(result[0],sumoUrl)
print("txt file name is:", result[0])
print("csv file name is:", result[1])
ctx = authneticateToSharepoint(kvUrl, certName, sharepointUrl)
fileDownload = downloadFile(ctx, sharepointFilePath)
updateExcel(fileDownload, result[1])
uploadFile (ctx, fileDownload, sharepointDir)
sendEmail(result[1])

'''